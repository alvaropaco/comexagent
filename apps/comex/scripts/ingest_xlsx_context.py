from __future__ import annotations

import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

import httpx
from openpyxl import load_workbook


def _iso_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _iter_sheet_rows(
    *,
    ws,
    max_rows: int | None = None,
) -> Iterable[List[Any]]:
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if max_rows is not None and i > max_rows:
            return
        yield list(row)


def _normalize_cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(v)
    return str(v).strip()


def _chunk_rows(rows: List[List[Any]], chunk_size: int) -> Iterable[Tuple[int, int, List[List[Any]]]]:
    start_idx = 0
    while start_idx < len(rows):
        end_idx = min(len(rows), start_idx + chunk_size)
        yield (start_idx, end_idx, rows[start_idx:end_idx])
        start_idx = end_idx


def _rows_to_text(*, file_name: str, sheet: str, header: List[str], start_row: int, rows: List[List[Any]]) -> str:
    lines: List[str] = []
    lines.append(f"Context dataset: {file_name}")
    lines.append(f"Sheet: {sheet}")
    lines.append(f"Rows: {start_row}-{start_row + len(rows) - 1}")
    lines.append("Columns: " + ", ".join(header))
    lines.append(f"IngestedAt: {_iso_now()}")
    lines.append("")

    for r in rows:
        kv = []
        for k, v in zip(header, r):
            sv = _normalize_cell(v)
            if sv:
                kv.append(f"{k}={sv}")
        if kv:
            lines.append(" | ".join(kv))
    return "\n".join(lines).strip()


def ingest_xlsx(
    *,
    xlsx_path: Path,
    core_data_api_url: str,
    vector_store_token: str,
    chunk_size: int = 200,
    max_rows_per_sheet: int | None = None,
    type_tag: str = "context_xlsx",
) -> Dict[str, Any]:
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    base_url = core_data_api_url.rstrip("/")

    sent = 0
    skipped = 0
    errors: List[str] = []

    with httpx.Client(timeout=60.0) as client:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_rows = list(_iter_sheet_rows(ws=ws, max_rows=max_rows_per_sheet))
            if not all_rows:
                continue
            header_raw = all_rows[0]
            header = [(_normalize_cell(x) or f"col_{i+1}") for i, x in enumerate(header_raw)]
            data_rows = all_rows[1:]
            if not data_rows:
                continue

            commodity = _detect_commodity(xlsx_path.name, sheet_name, header, data_rows[:25])
            origin = _detect_origin(header, data_rows[:25])
            date = _extract_date_from_filename(xlsx_path.name)

            for start_idx, end_idx, chunk in _chunk_rows(data_rows, chunk_size=chunk_size):
                text = _rows_to_text(
                    file_name=xlsx_path.name,
                    sheet=sheet_name,
                    header=header,
                    start_row=2 + start_idx,
                    rows=chunk,
                )
                if not text:
                    skipped += 1
                    continue

                metadata = {
                    "type": type_tag,
                    "source": "local_file",
                    "path": str(xlsx_path),
                    "file": xlsx_path.name,
                    "sheet": sheet_name,
                    "row_start": 2 + start_idx,
                    "row_end": 1 + end_idx,
                    "ingestedAt": _iso_now(),
                    "commodity": commodity,
                    "topic": "logistics",
                    "origin": origin,
                    "date": date,
                }

                try:
                    r = client.post(
                        f"{base_url}/vector/store",
                        headers={"x-vector-store-token": vector_store_token},
                        json={"text": text, "metadata": metadata},
                    )
                    r.raise_for_status()
                    sent += 1
                except Exception as e:
                    errors.append(f"{sheet_name}:{start_idx}-{end_idx}:{e.__class__.__name__}")

    return {"ok": not errors, "sent": sent, "skipped": skipped, "errors": errors}


def _extract_date_from_filename(name: str) -> str | None:
    m = re.search(r"DT(\d{8})", name)
    if not m:
        return None
    s = m.group(1)
    return f"{s[0:4]}-{s[4:6]}-{s[6:8]}T00:00:00+00:00"


def _detect_commodity(file_name: str, sheet: str, header: List[str], rows: List[List[Any]]) -> str | None:
    blob = " ".join([file_name, sheet] + header).lower()
    if "cafe" in blob or "coffee" in blob:
        return "coffee"
    if "pimenta" in blob or "pepper" in blob:
        return "pepper"

    for r in rows:
        for v in r:
            sv = _normalize_cell(v).lower()
            if "cafe" in sv or "coffee" in sv:
                return "coffee"
            if "pimenta" in sv or "pepper" in sv:
                return "pepper"
    return None


def _detect_origin(header: List[str], rows: List[List[Any]]) -> str | None:
    header_l = [h.lower() for h in header]
    for key in ("origin", "origem", "port", "porto", "mun", "municip", "uf", "estado"):
        for idx, h in enumerate(header_l):
            if key in h:
                for r in rows:
                    if idx < len(r):
                        val = _normalize_cell(r[idx])
                        if val:
                            return val
    return None


def main() -> None:
    xlsx = os.getenv("XLSX_PATH")
    core_url = os.getenv("CORE_DATA_API_URL")
    token = os.getenv("VECTOR_STORE_TOKEN")
    chunk_size = int(os.getenv("CHUNK_SIZE", "200"))
    max_rows = os.getenv("MAX_ROWS_PER_SHEET")
    max_rows_val = int(max_rows) if max_rows else None

    if not xlsx or not core_url or not token:
        raise SystemExit(
            "Missing env vars. Set XLSX_PATH, CORE_DATA_API_URL, VECTOR_STORE_TOKEN. "
            "Optional: CHUNK_SIZE, MAX_ROWS_PER_SHEET."
        )

    res = ingest_xlsx(
        xlsx_path=Path(xlsx),
        core_data_api_url=core_url,
        vector_store_token=token,
        chunk_size=chunk_size,
        max_rows_per_sheet=max_rows_val,
    )
    print(json.dumps(res, indent=2))


if __name__ == "__main__":
    main()
